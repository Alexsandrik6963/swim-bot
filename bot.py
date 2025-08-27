import os
import asyncio
from aiogram import Bot, Dispatcher, types, F
from aiogram.types import Message, CallbackQuery
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from dotenv import load_dotenv
from datetime import datetime, timedelta
import pytz
from openpyxl import Workbook, load_workbook

# Загружаем .env
load_dotenv()
TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID", 191598071))  # твой ID

bot = Bot(
    token=TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.HTML)
)
dp = Dispatcher()

# Часовой пояс Москва
moscow_tz = pytz.timezone("Europe/Moscow")

# Файл для записей
EXCEL_FILE = "bookings.xlsx"

# --- СОХРАНЕНИЕ ЗАПИСЕЙ В EXCEL ---
def save_booking(date_str, time_str, user, status="Записан"):
    """Сохраняет запись в Excel"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата", "Время", "UserID", "Username", "Статус"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        date_str,
        time_str,
        user.id,
        user.username or "",
        status
    ])
    wb.save(EXCEL_FILE)


# Рабочие дни (0=понедельник ... 6=воскресенье)
WORK_DAYS = [0, 1, 2, 3, 5]  # Пн-Чт и Сб

# --- КЛАВИАТУРЫ ---

def main_menu():
    kb = InlineKeyboardBuilder()
    kb.button(text="📅 Записаться на тренировку", callback_data="choose_day")
    kb.button(text="❌ Отмена записи", callback_data="cancel")
    return kb.as_markup()

def days_keyboard():
    kb = InlineKeyboardBuilder()
    today = datetime.now(moscow_tz).date()
    for i in range(14):  # ближайшие 2 недели
        d = today + timedelta(days=i)
        if d.weekday() in WORK_DAYS:
            kb.button(text=d.strftime("%a %d.%m"), callback_data=f"day_{d}")
    kb.button(text="⬅️ Назад", callback_data="back_menu")
    return kb.adjust(2).as_markup()

def times_keyboard(selected_date):
    kb = InlineKeyboardBuilder()
    date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()
    weekday = date_obj.weekday()

    slots = []

    if weekday in [0, 1, 2, 3]:  # Пн–Чт
        slots = [
            "07:00","08:00","09:00","10:00","11:00","12:00",
            "13:00","14:00","15:00",
            "18:30","19:00","20:00","21:00"
        ]
    elif weekday == 5:  # Суббота
        slots = [
            "07:00","08:00","09:00","10:00","11:00","12:00",
            "13:00","14:00","15:00"
        ]

    for s in slots:
        kb.button(text=s, callback_data=f"time_{selected_date}_{s}")

    kb.button(text="⬅️ Назад", callback_data="choose_day")
    return kb.adjust(4).as_markup()

# --- ХЕНДЛЕРЫ ---

@dp.message(F.text == "/start")
async def cmd_start(message: Message):
    await message.answer(
        "Привет! 👋 Я бот для записи на тренировки по плаванию.\nВыбери действие:",
        reply_markup=main_menu()
    )

@dp.callback_query(F.data == "back_menu")
async def back_menu(cb: CallbackQuery):
    await cb.message.edit_text("Главное меню:", reply_markup=main_menu())
    await cb.answer()

@dp.callback_query(F.data == "choose_day")
async def choose_day(cb: CallbackQuery):
    await cb.message.edit_text("Выбери день:", reply_markup=days_keyboard())
    await cb.answer()

@dp.callback_query(F.data.startswith("day_"))
async def choose_time(cb: CallbackQuery):
    date_str = cb.data.split("_")[1]
    await cb.message.edit_text(
        f"Выбери время для <b>{date_str}</b>:",
        reply_markup=times_keyboard(date_str)
    )
    await cb.answer()

@dp.callback_query(F.data.startswith("time_"))
async def confirm_booking(cb: CallbackQuery):
    _, date_str, time_str = cb.data.split("_")
    booking_text = f"✅ Запись подтверждена!\n📅 Дата: {date_str}\n⏰ Время: {time_str}"

    await cb.message.edit_text(booking_text, reply_markup=main_menu())
    await cb.answer("Запись успешно создана!")

    # Сохраняем запись в Excel
    save_booking(date_str, time_str, cb.from_user, status="Записан")

    # Отправляем админу в личку
    try:
        await bot.send_message(
            ADMIN_ID,
            f"📌 Новая запись!\n👤 Пользователь: @{cb.from_user.username or cb.from_user.id}\n📅 Дата: {date_str}\n⏰ Время: {time_str}"
        )
    except Exception as e:
        print(f"Ошибка при отправке админу: {e}")

@dp.callback_query(F.data == "cancel")
async def cancel(cb: CallbackQuery):
    await cb.message.edit_text("❌ Ваша запись отменена.", reply_markup=main_menu())
    await cb.answer("Отмена выполнена")

    # Сохраняем отмену в Excel
    save_booking("-", "-", cb.from_user, status="Отменил")

    # Уведомляем админа
    try:
        await bot.send_message(
            ADMIN_ID,
            f"⚠️ Пользователь @{cb.from_user.username or cb.from_user.id} отменил запись."
        )
    except Exception as e:
        print(f"Ошибка при отправке админу: {e}")

# --- ЗАПУСК ---
async def main():
    # Сообщение админу при запуске
    try:
        await bot.send_message(ADMIN_ID, "✅ Бот запущен и работает.")
    except Exception as e:
        print(f"Не удалось отправить админу сообщение при старте: {e}")

    print("Бот запущен...")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
